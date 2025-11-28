"""
Parser service for processing Buz inventory and pricing Excel exports.

Parses Excel files exported from Buz and extracts inventory items and
pricing coefficients into structured data.

Buz export format quirks:
- Inventory files: Row 1 is title (e.g., "Glideshift Blinds - 28/11/2025"),
  headers are on row 2
- Pricing files: Headers are on row 1
- Both have an Operation column (A=Add, E=Edit, D=Delete) for import
- Each sheet represents an inventory group (sheet name = group code)
- 'Help' and 'Sheet1' sheets should be skipped
"""
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Sheets to skip when parsing
SKIP_SHEETS = {'help', 'sheet1'}


class BuzExcelParser:
    """
    Parser for Buz Excel exports.

    Handles both inventory items and pricing coefficients exports.
    The parser is flexible to handle varying column structures.
    """

    # Column name mappings for Buz inventory items export
    # Buz exports use "Code*" and "Description*" with asterisks
    INVENTORY_COLUMN_MAP = {
        # Primary key (Buz internal)
        'pkid': 'pk_id',

        # Code columns - Buz uses "Code*" with asterisk
        'code*': 'item_code',
        'code': 'item_code',
        'item code': 'item_code',
        'inventory group code': 'group_code',
        'inventorygroupcode': 'group_code',
        'group code': 'group_code',

        # Description columns - Buz uses "Description*" with asterisk
        'description*': 'description',
        'description': 'description',
        'descnpart1 (material)': 'material',
        'descnpart2 (material types)': 'material_type',
        'descnpart3 (colour)': 'colour',

        # Grid codes
        'price grid code': 'price_grid_code',
        'cost grid code': 'cost_grid_code',
        'discount group code': 'discount_group_code',

        # Pricing
        'last purchase price': 'last_purchase_price',
        'standard cost': 'standard_cost',
        'tax rate': 'tax_rate',

        # Units and quantities
        'units purchase': 'units_purchase',
        'min qty': 'min_qty',
        'max qty': 'max_qty',
        'reorder multiplier': 'reorder_multiplier',
        'stocking multiplier': 'stocking_multiplier',
        'units stock': 'units_stock',
        'selling multiplier': 'selling_multiplier',
        'units sell': 'units_sell',

        # ForeX (foreign exchange)
        'forex code': 'forex_code',
        'last purchase forex': 'last_purchase_forex',

        # Lead times
        'purchasing lead days': 'purchasing_lead_days',

        # Product info
        'cost method': 'cost_method',
        'product size': 'product_size',
        'product type': 'product_type',

        # Supplier columns
        'supplier': 'supplier_name',
        'supplier code': 'supplier_code',
        'supplier product code': 'supplier_product_code',
        'supplier product description': 'supplier_product_description',

        # Dimensions
        'length': 'length',
        'maximum width': 'maximum_width',

        # Production times
        'extra time to produce': 'extra_time_to_produce',
        'extra time to fit': 'extra_time_to_fit',

        # Custom variables
        'custom var 1 (packsize)': 'pack_size',
        'custom var 2 (packopt)': 'pack_option',
        'custom var 3 (packtype)': 'pack_type',

        # Other
        'warning': 'warning',
        'rptcat': 'report_category',
        'last edit date': 'last_edit_date',

        # Status columns
        'active': 'is_active',
        'is current': 'is_active',
        'iscurrent': 'is_active',

        # Operation column (A=Add, E=Edit, D=Delete)
        'operation': 'operation',
    }

    # Column name mappings for Buz pricing export
    # Note: IsNotCurrent is inverted (True = inactive)
    PRICING_COLUMN_MAP = {
        # Primary key (Buz internal)
        'pkid': 'pk_id',

        # Item reference columns
        'inventory code': 'item_code',
        'inventorycode': 'item_code',
        'code': 'item_code',

        # Description
        'description': 'description',

        # Customer price group
        'customer price group code': 'price_group_code',
        'customerpricegroupcode': 'price_group_code',
        'price group code': 'price_group_code',
        'price group': 'price_group_code',

        # Effective date
        'date from': 'effective_date',
        'datefrom': 'effective_date',
        'effective date': 'effective_date',

        # Sell prices - various calculation methods
        'sell each': 'sell_each',
        'selleach': 'sell_each',
        'selllmwide': 'sell_lm_wide',
        'selllmheight': 'sell_lm_height',
        'selllmdepth': 'sell_lm_depth',
        'sellsqm': 'sell_sqm',
        'sellpercentageonmain': 'sell_percentage_on_main',
        'sellminimum': 'sell_minimum',

        # Cost prices - various calculation methods
        'costeach': 'cost_each',
        'costlmwide': 'cost_lm_wide',
        'costlmheight': 'cost_lm_height',
        'costlmdepth': 'cost_lm_depth',
        'costsqm': 'cost_sqm',
        'costpercentageonmain': 'cost_percentage_on_main',
        'costminimum': 'cost_minimum',

        # Install costs
        'installcosteach': 'install_cost_each',
        'installcostlmwidth': 'install_cost_lm_width',
        'installcost height': 'install_cost_height',
        'installcostdepth': 'install_cost_depth',
        'installcostsqm': 'install_cost_sqm',
        'installcostpercentageofmain': 'install_cost_percentage_of_main',
        'installcostminimum': 'install_cost_minimum',

        # Install sell prices
        'installselleach': 'install_sell_each',
        'installsellminimum': 'install_sell_minimum',
        'installselllmwide': 'install_sell_lm_wide',
        'installsellsqm': 'install_sell_sqm',
        'installsellheight': 'install_sell_height',
        'installselldepth': 'install_sell_depth',
        'installsellpercentageofmain': 'install_sell_percentage_of_main',

        # Supplier
        'supplier code': 'supplier_code',
        'supplier descn': 'supplier_description',

        # Sort order
        'sortorder': 'sort_order',
        'sort order': 'sort_order',

        # Status - NOTE: IsNotCurrent is inverted (True = inactive)
        'isnotcurrent': 'is_not_current',
        'is not current': 'is_not_current',

        # Operation column (A=Add for pricing)
        'operation': 'operation',
    }

    def parse_inventory_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse an inventory items Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dict with items list and metadata
        """
        logger.info(f"Parsing inventory file: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            items = []
            groups = {}

            for sheet_name in workbook.sheetnames:
                # Skip Help and Sheet1 sheets
                if sheet_name.lower() in SKIP_SHEETS:
                    logger.debug(f"Skipping sheet: {sheet_name}")
                    continue

                sheet = workbook[sheet_name]
                sheet_items = self._parse_inventory_sheet(sheet, sheet_name)
                items.extend(sheet_items)

                # Track groups - use sheet name as group code
                for item in sheet_items:
                    group_code = item.get('group_code') or sheet_name
                    if group_code not in groups:
                        groups[group_code] = {
                            'group_code': group_code,
                            'group_name': group_code,
                            'item_count': 0
                        }
                    groups[group_code]['item_count'] += 1

            workbook.close()

            logger.info(f"Parsed {len(items)} inventory items from {len(groups)} groups")

            return {
                'success': True,
                'items': items,
                'groups': list(groups.values()),
                'total_items': len(items),
                'total_groups': len(groups)
            }

        except Exception as e:
            logger.exception(f"Error parsing inventory file: {e}")
            return {
                'success': False,
                'error': str(e),
                'items': [],
                'groups': []
            }

    def parse_pricing_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse a pricing coefficients Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dict with coefficients list and metadata
        """
        logger.info(f"Parsing pricing file: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            coefficients = []
            groups = {}

            for sheet_name in workbook.sheetnames:
                # Skip Help and Sheet1 sheets
                if sheet_name.lower() in SKIP_SHEETS:
                    logger.debug(f"Skipping sheet: {sheet_name}")
                    continue

                sheet = workbook[sheet_name]
                sheet_pricing = self._parse_pricing_sheet(sheet, sheet_name)
                coefficients.extend(sheet_pricing)

                # Track groups - use sheet name as group code
                for pricing in sheet_pricing:
                    group_code = pricing.get('group_code') or sheet_name
                    if group_code not in groups:
                        groups[group_code] = {
                            'group_code': group_code,
                            'group_name': group_code,
                            'coefficient_count': 0
                        }
                    groups[group_code]['coefficient_count'] += 1

            workbook.close()

            logger.info(f"Parsed {len(coefficients)} pricing coefficients from {len(groups)} groups")

            return {
                'success': True,
                'coefficients': coefficients,
                'groups': list(groups.values()),
                'total_coefficients': len(coefficients),
                'total_groups': len(groups)
            }

        except Exception as e:
            logger.exception(f"Error parsing pricing file: {e}")
            return {
                'success': False,
                'error': str(e),
                'coefficients': [],
                'groups': []
            }

    def _parse_inventory_sheet(
        self,
        sheet: Worksheet,
        default_group: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a single sheet of inventory items.

        Args:
            sheet: openpyxl worksheet
            default_group: Default group code if not in data

        Returns:
            List of inventory item dicts
        """
        items = []
        header_row = None
        column_map = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(row):
                continue

            # Find header row
            if header_row is None:
                if self._looks_like_header(row, self.INVENTORY_COLUMN_MAP):
                    header_row = row_idx
                    column_map = self._build_column_map(row, self.INVENTORY_COLUMN_MAP)
                    logger.debug(f"Found inventory header at row {row_idx}: {column_map}")
                continue

            # Parse data row
            item = self._parse_inventory_row(row, column_map, default_group)
            if item:
                items.append(item)

        return items

    def _parse_pricing_sheet(
        self,
        sheet: Worksheet,
        default_group: str
    ) -> List[Dict[str, Any]]:
        """
        Parse a single sheet of pricing coefficients.

        Args:
            sheet: openpyxl worksheet
            default_group: Default group code if not in data

        Returns:
            List of pricing coefficient dicts
        """
        coefficients = []
        header_row = None
        column_map = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(row):
                continue

            # Find header row
            if header_row is None:
                if self._looks_like_header(row, self.PRICING_COLUMN_MAP):
                    header_row = row_idx
                    column_map = self._build_column_map(row, self.PRICING_COLUMN_MAP)
                    logger.debug(f"Found pricing header at row {row_idx}: {column_map}")
                continue

            # Parse data row
            coeff = self._parse_pricing_row(row, column_map, default_group)
            if coeff:
                coefficients.append(coeff)

        return coefficients

    def _looks_like_header(self, row: tuple, mapping: Dict[str, str]) -> bool:
        """Check if a row looks like a header based on known column names."""
        if not row:
            return False

        matches = 0
        for cell in row:
            if cell and isinstance(cell, str):
                normalized = cell.lower().strip()
                if normalized in mapping:
                    matches += 1

        # Consider it a header if we match at least 2 columns
        return matches >= 2

    def _build_column_map(
        self,
        header_row: tuple,
        mapping: Dict[str, str]
    ) -> Dict[int, str]:
        """
        Build a mapping of column index to field name.

        Args:
            header_row: Tuple of header values
            mapping: Dict of header name -> field name

        Returns:
            Dict of column_index -> field_name
        """
        column_map = {}
        for idx, cell in enumerate(header_row):
            if cell and isinstance(cell, str):
                normalized = cell.lower().strip()
                if normalized in mapping:
                    column_map[idx] = mapping[normalized]
        return column_map

    def _parse_inventory_row(
        self,
        row: tuple,
        column_map: Dict[int, str],
        default_group: str
    ) -> Optional[Dict[str, Any]]:
        """Parse a single inventory item row."""
        item = {
            'group_code': default_group,
            'pk_id': None,
            'item_code': '',
            'description': '',
            'description_part1': '',
            'description_part2': '',
            'description_part3': '',
            'price_grid_code': '',
            'cost_grid_code': '',
            'unit_of_measure': '',
            'is_active': True,
            'supplier_code': '',
            'supplier_name': '',
            'sort_order': 0,
            'operation': '',
            'extra_data': {}
        }

        unmapped = {}

        for idx, value in enumerate(row):
            if idx in column_map:
                field = column_map[idx]
                item[field] = self._convert_value(value, field)
            elif value is not None:
                # Store unmapped columns in extra_data
                unmapped[f'col_{idx}'] = value

        if unmapped:
            item['extra_data'] = unmapped

        # Item must have at least a code
        if not item['item_code']:
            return None

        return item

    def _parse_pricing_row(
        self,
        row: tuple,
        column_map: Dict[int, str],
        default_group: str
    ) -> Optional[Dict[str, Any]]:
        """Parse a single pricing row."""
        pricing = {
            'group_code': default_group,
            'pk_id': None,
            'item_code': '',
            'description': '',
            'price_group_code': '',
            'effective_date': None,
            'is_active': True,
            # Sell prices - various calculation methods
            'sell_each': 0.0,
            'sell_lm_wide': 0.0,
            'sell_lm_height': 0.0,
            'sell_lm_depth': 0.0,
            'sell_sqm': 0.0,
            'sell_percentage_on_main': 0.0,
            'sell_minimum': 0.0,
            # Cost prices
            'cost_each': 0.0,
            'cost_lm_wide': 0.0,
            'cost_lm_height': 0.0,
            'cost_lm_depth': 0.0,
            'cost_sqm': 0.0,
            'cost_percentage_on_main': 0.0,
            'cost_minimum': 0.0,
            # Install costs
            'install_cost_each': 0.0,
            'install_cost_lm_width': 0.0,
            'install_cost_height': 0.0,
            'install_cost_depth': 0.0,
            'install_cost_sqm': 0.0,
            'install_cost_percentage_of_main': 0.0,
            'install_cost_minimum': 0.0,
            # Install sell prices
            'install_sell_each': 0.0,
            'install_sell_minimum': 0.0,
            'install_sell_lm_wide': 0.0,
            'install_sell_sqm': 0.0,
            'install_sell_height': 0.0,
            'install_sell_depth': 0.0,
            'install_sell_percentage_of_main': 0.0,
            # Supplier
            'supplier_code': '',
            'supplier_description': '',
            'sort_order': 0,
            'operation': '',
            'extra_data': {}
        }

        unmapped = {}

        for idx, value in enumerate(row):
            if idx in column_map:
                field = column_map[idx]
                pricing[field] = self._convert_value(value, field)
            elif value is not None:
                # Store unmapped columns in extra_data
                unmapped[f'col_{idx}'] = value

        if unmapped:
            pricing['extra_data'] = unmapped

        # Handle IsNotCurrent inversion (True = inactive, so invert for is_active)
        if 'is_not_current' in pricing:
            is_not_current = pricing.pop('is_not_current')
            # Invert: if is_not_current is True, is_active should be False
            if isinstance(is_not_current, bool):
                pricing['is_active'] = not is_not_current
            elif isinstance(is_not_current, str):
                pricing['is_active'] = is_not_current.lower() not in ('yes', 'true', '1', 'y')
            elif isinstance(is_not_current, (int, float)):
                pricing['is_active'] = not bool(is_not_current)

        # Pricing must have at least an item code
        if not pricing['item_code']:
            return None

        return pricing

    # Fields that should be treated as numeric (floats)
    NUMERIC_FIELDS = {
        # Inventory numeric fields
        'last_purchase_price', 'standard_cost', 'tax_rate',
        'min_qty', 'max_qty', 'reorder_multiplier',
        'stocking_multiplier', 'selling_multiplier',
        'last_purchase_forex', 'length', 'maximum_width',
        'extra_time_to_produce', 'extra_time_to_fit',
        # Pricing numeric fields - sell
        'sell_each', 'sell_lm_wide', 'sell_lm_height', 'sell_lm_depth',
        'sell_sqm', 'sell_percentage_on_main', 'sell_minimum',
        # Pricing numeric fields - cost
        'cost_each', 'cost_lm_wide', 'cost_lm_height', 'cost_lm_depth',
        'cost_sqm', 'cost_percentage_on_main', 'cost_minimum',
        # Pricing numeric fields - install cost
        'install_cost_each', 'install_cost_lm_width', 'install_cost_height',
        'install_cost_depth', 'install_cost_sqm',
        'install_cost_percentage_of_main', 'install_cost_minimum',
        # Pricing numeric fields - install sell
        'install_sell_each', 'install_sell_minimum', 'install_sell_lm_wide',
        'install_sell_sqm', 'install_sell_height', 'install_sell_depth',
        'install_sell_percentage_of_main',
    }

    def _convert_value(self, value: Any, field: str) -> Any:
        """Convert a cell value to the appropriate type for a field."""
        from datetime import datetime

        if value is None:
            # Return appropriate default based on field type
            if field in ('is_active', 'is_not_current'):
                return False if field == 'is_not_current' else True
            elif field in self.NUMERIC_FIELDS:
                return 0.0
            elif field in ('sort_order', 'pk_id', 'purchasing_lead_days'):
                return 0 if field != 'pk_id' else None
            elif field in ('effective_date', 'last_edit_date'):
                return None
            return ''

        # Boolean fields
        if field in ('is_active', 'is_not_current'):
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return bool(value)
            if isinstance(value, str):
                return value.lower() in ('yes', 'true', '1', 'y', 'active', 'current')
            return False

        # Date fields
        if field in ('effective_date', 'last_edit_date'):
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d')
            if isinstance(value, str):
                return value.strip()
            return str(value) if value else None

        # Numeric (float) fields
        if field in self.NUMERIC_FIELDS:
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0

        # Integer fields
        if field in ('sort_order', 'pk_id', 'purchasing_lead_days'):
            try:
                return int(value)
            except (ValueError, TypeError):
                return 0 if field != 'pk_id' else None

        # String fields
        return str(value).strip() if value else ''


# Singleton instance
parser = BuzExcelParser()
