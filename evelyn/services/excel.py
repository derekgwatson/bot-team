"""Excel processing service for Evelyn."""
import io
import logging
from typing import BinaryIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for processing Excel workbooks."""

    def get_sheet_names(self, file_data: BinaryIO) -> list[str]:
        """
        Get list of sheet names from an Excel workbook.

        Args:
            file_data: File-like object containing Excel data

        Returns:
            List of sheet names
        """
        wb = load_workbook(file_data, read_only=True, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names

    def process_workbook(
        self,
        file_data: BinaryIO,
        sheets_to_keep: list[str],
        original_filename: str = "workbook.xlsx"
    ) -> tuple[io.BytesIO, str]:
        """
        Process an Excel workbook: keep only specified sheets and convert to values.

        Args:
            file_data: File-like object containing Excel data
            sheets_to_keep: List of sheet names to keep
            original_filename: Original filename for generating output name

        Returns:
            Tuple of (BytesIO with processed workbook, suggested filename)
        """
        # Load workbook with data_only=False first to read formulas
        # We'll manually convert to values
        wb = load_workbook(file_data, data_only=False)

        all_sheets = wb.sheetnames
        logger.info(f"Processing workbook with sheets: {all_sheets}")
        logger.info(f"Keeping sheets: {sheets_to_keep}")

        # Validate requested sheets exist
        for sheet_name in sheets_to_keep:
            if sheet_name not in all_sheets:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        # Convert all cells to values (remove formulas)
        for sheet_name in sheets_to_keep:
            ws = wb[sheet_name]
            self._convert_sheet_to_values(ws)

        # Delete sheets we don't want to keep
        sheets_to_delete = [s for s in all_sheets if s not in sheets_to_keep]
        for sheet_name in sheets_to_delete:
            del wb[sheet_name]
            logger.info(f"Deleted sheet: {sheet_name}")

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        # Generate output filename
        base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
        output_filename = f"{base_name}_processed.xlsx"

        return output, output_filename

    def _convert_sheet_to_values(self, ws) -> None:
        """
        Convert all formulas in a worksheet to their calculated values.

        Note: This reads the formula cells and replaces them with values.
        For cells with formulas, we need to load with data_only=True separately
        to get the cached values.
        """
        # We need to reload the workbook with data_only=True to get values
        # But since we're working in memory, we'll take a different approach:
        # We iterate through all cells and replace formula cells with their value

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                    # This is a formula - we need the cached value
                    # openpyxl stores the cached value in cell.value when data_only=True
                    # Since we loaded with data_only=False, we need to handle this differently
                    pass  # Will be handled by the dual-load approach below


# Enhanced version that properly handles formula-to-value conversion
class ExcelServiceV2:
    """Enhanced Excel service with proper formula-to-value conversion."""

    def get_sheet_names(self, file_data: BinaryIO) -> list[str]:
        """Get list of sheet names from an Excel workbook."""
        wb = load_workbook(file_data, read_only=True, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names

    def process_workbook(
        self,
        file_data: BinaryIO,
        sheets_to_keep: list[str],
        original_filename: str = "workbook.xlsx"
    ) -> tuple[io.BytesIO, str]:
        """
        Process an Excel workbook: keep only specified sheets and convert to values.

        This uses a dual-load approach:
        1. Load with data_only=True to get calculated values
        2. Load with data_only=False to get formatting and structure
        3. Copy values from (1) to (2), then remove unwanted sheets
        """
        # First pass: load with data_only=True to get calculated values
        file_data.seek(0)
        wb_values = load_workbook(file_data, data_only=True)

        # Second pass: load with data_only=False to preserve formatting
        file_data.seek(0)
        wb_format = load_workbook(file_data, data_only=False)

        all_sheets = wb_format.sheetnames
        logger.info(f"Processing workbook with sheets: {all_sheets}")
        logger.info(f"Keeping sheets: {sheets_to_keep}")

        # Validate requested sheets exist
        for sheet_name in sheets_to_keep:
            if sheet_name not in all_sheets:
                wb_values.close()
                wb_format.close()
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        # Copy calculated values to the formatted workbook for sheets we're keeping
        for sheet_name in sheets_to_keep:
            ws_values = wb_values[sheet_name]
            ws_format = wb_format[sheet_name]

            # Get dimensions
            max_row = ws_format.max_row or 1
            max_col = ws_format.max_column or 1

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell_format = ws_format.cell(row=row, column=col)
                    cell_values = ws_values.cell(row=row, column=col)

                    # Replace the value (formula becomes calculated value)
                    # Keep the formatting from ws_format
                    if cell_format.value is not None:
                        if isinstance(cell_format.value, str) and cell_format.value.startswith('='):
                            # It's a formula - replace with calculated value
                            cell_format.value = cell_values.value
                        # Non-formula values are already correct

        wb_values.close()

        # Delete sheets we don't want to keep
        sheets_to_delete = [s for s in all_sheets if s not in sheets_to_keep]
        for sheet_name in sheets_to_delete:
            del wb_format[sheet_name]
            logger.info(f"Deleted sheet: {sheet_name}")

        # Save to BytesIO
        output = io.BytesIO()
        wb_format.save(output)
        wb_format.close()
        output.seek(0)

        # Generate output filename
        base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
        output_filename = f"{base_name}_processed.xlsx"

        return output, output_filename


# Use the enhanced version
excel_service = ExcelServiceV2()
