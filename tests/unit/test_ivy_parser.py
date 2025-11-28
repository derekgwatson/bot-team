"""
Unit tests for Ivy Excel parser service.

Requires openpyxl to be installed. Tests are skipped if openpyxl is not available.
"""

import os
import sys
import pytest
import importlib.util
from pathlib import Path

# Check if openpyxl is available
try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Skip all tests in this module if openpyxl is not available
pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")

# Add project root to path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Set test environment
os.environ['TESTING'] = '1'
os.environ['SKIP_ENV_VALIDATION'] = '1'
os.environ['FLASK_SECRET_KEY'] = 'test-secret-key'
os.environ['BUZ_ORGS'] = ''

# Add ivy to path
sys.path.insert(0, str(project_root / 'ivy'))

# Only import parser if openpyxl is available
BuzExcelParser = None
if HAS_OPENPYXL:
    # Import parser using importlib
    module_path = project_root / 'ivy' / 'services' / 'parser_service.py'
    spec = importlib.util.spec_from_file_location('ivy_parser_service', module_path)
    parser_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(parser_module)
    BuzExcelParser = parser_module.BuzExcelParser


@pytest.fixture
def parser():
    """Create a parser instance for testing."""
    return BuzExcelParser()


@pytest.fixture
def sample_inventory_excel(tmp_path):
    """Create a sample inventory Excel file for testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROLL"

    # Headers
    ws.append(['Item Code', 'Item Name', 'Description', 'Is Current', 'Cost Price', 'Sell Price'])

    # Data rows
    ws.append(['ROLL001', 'Standard Roller', 'A standard roller blind', 'Yes', 50.00, 99.00])
    ws.append(['ROLL002', 'Premium Roller', 'A premium roller blind', 'Yes', 75.00, 149.00])
    ws.append(['ROLL003', 'Budget Roller', 'Budget option', 'No', 30.00, 49.00])

    file_path = tmp_path / "inventory_test.xlsx"
    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def sample_pricing_excel(tmp_path):
    """Create a sample pricing Excel file for testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROLL"

    # Headers
    ws.append(['Code', 'Name', 'Type', 'Is Current', 'Value'])

    # Data rows
    ws.append(['WIDTH_MULT', 'Width Multiplier', 'Multiplier', 'Yes', 1.25])
    ws.append(['HEIGHT_MULT', 'Height Multiplier', 'Multiplier', 'Yes', 1.15])
    ws.append(['OLD_COEFF', 'Old Coefficient', 'Deprecated', 'No', 0.5])

    file_path = tmp_path / "pricing_test.xlsx"
    wb.save(file_path)
    return str(file_path)


@pytest.mark.unit
@pytest.mark.ivy
class TestInventoryParsing:
    """Test inventory Excel parsing."""

    def test_parse_inventory_file(self, parser, sample_inventory_excel):
        """Test parsing a valid inventory file."""
        result = parser.parse_inventory_file(sample_inventory_excel)

        assert result['success'] is True
        assert result['total_items'] == 3
        assert result['total_groups'] == 1

        items = result['items']
        assert len(items) == 3

        # Check first item
        item1 = next(i for i in items if i['item_code'] == 'ROLL001')
        assert item1['item_name'] == 'Standard Roller'
        assert item1['description'] == 'A standard roller blind'
        assert item1['is_active'] is True
        assert item1['cost_price'] == 50.00
        assert item1['sell_price'] == 99.00

        # Check inactive item
        item3 = next(i for i in items if i['item_code'] == 'ROLL003')
        assert item3['is_active'] is False

    def test_parse_inventory_groups(self, parser, sample_inventory_excel):
        """Test that groups are extracted correctly."""
        result = parser.parse_inventory_file(sample_inventory_excel)

        groups = result['groups']
        assert len(groups) == 1

        group = groups[0]
        assert group['group_code'] == 'ROLL'
        assert group['item_count'] == 3

    def test_parse_inventory_nonexistent_file(self, parser):
        """Test parsing a file that doesn't exist."""
        result = parser.parse_inventory_file('/nonexistent/file.xlsx')

        assert result['success'] is False
        assert 'error' in result


@pytest.mark.unit
@pytest.mark.ivy
class TestPricingParsing:
    """Test pricing Excel parsing."""

    def test_parse_pricing_file(self, parser, sample_pricing_excel):
        """Test parsing a valid pricing file."""
        result = parser.parse_pricing_file(sample_pricing_excel)

        assert result['success'] is True
        assert result['total_coefficients'] == 3
        assert result['total_groups'] == 1

        coefficients = result['coefficients']
        assert len(coefficients) == 3

        # Check first coefficient
        coeff1 = next(c for c in coefficients if c['coefficient_code'] == 'WIDTH_MULT')
        assert coeff1['coefficient_name'] == 'Width Multiplier'
        assert coeff1['coefficient_type'] == 'Multiplier'
        assert coeff1['is_active'] is True
        assert coeff1['base_value'] == 1.25

        # Check inactive coefficient
        coeff3 = next(c for c in coefficients if c['coefficient_code'] == 'OLD_COEFF')
        assert coeff3['is_active'] is False


@pytest.mark.unit
@pytest.mark.ivy
class TestColumnMapping:
    """Test column name mapping flexibility."""

    def test_column_name_variations(self, parser, tmp_path):
        """Test that various column name formats are recognized."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        # Use different column name formats
        ws.append(['ItemCode', 'ItemName', 'IsCurrent', 'CostPrice'])
        ws.append(['TEST001', 'Test Item', 'True', 100])

        file_path = tmp_path / "variations.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))

        assert result['success'] is True
        assert len(result['items']) == 1

        item = result['items'][0]
        assert item['item_code'] == 'TEST001'
        assert item['item_name'] == 'Test Item'

    def test_active_status_variations(self, parser, tmp_path):
        """Test that various active status values are parsed correctly."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['Item Code', 'Item Name', 'Is Current'])
        ws.append(['ITEM1', 'Yes Item', 'Yes'])
        ws.append(['ITEM2', 'True Item', 'True'])
        ws.append(['ITEM3', 'One Item', '1'])
        ws.append(['ITEM4', 'No Item', 'No'])
        ws.append(['ITEM5', 'False Item', 'False'])
        ws.append(['ITEM6', 'Zero Item', '0'])

        file_path = tmp_path / "active_test.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))
        items = {i['item_code']: i for i in result['items']}

        assert items['ITEM1']['is_active'] is True
        assert items['ITEM2']['is_active'] is True
        assert items['ITEM3']['is_active'] is True
        assert items['ITEM4']['is_active'] is False
        assert items['ITEM5']['is_active'] is False
        assert items['ITEM6']['is_active'] is False


@pytest.mark.unit
@pytest.mark.ivy
class TestEmptyAndEdgeCases:
    """Test edge cases and empty data."""

    def test_empty_sheet(self, parser, tmp_path):
        """Test parsing an empty Excel file."""
        import openpyxl

        wb = openpyxl.Workbook()
        file_path = tmp_path / "empty.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))

        assert result['success'] is True
        assert result['total_items'] == 0

    def test_header_only(self, parser, tmp_path):
        """Test parsing a file with only headers."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Item Code', 'Item Name', 'Is Current'])

        file_path = tmp_path / "header_only.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))

        assert result['success'] is True
        assert result['total_items'] == 0

    def test_missing_required_fields(self, parser, tmp_path):
        """Test that rows without code or name are skipped."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['Item Code', 'Item Name', 'Description'])
        ws.append(['', '', 'No code or name'])  # Should be skipped
        ws.append(['VALID', 'Valid Item', 'Has both'])  # Should be included

        file_path = tmp_path / "missing_fields.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))

        assert result['success'] is True
        assert result['total_items'] == 1
        assert result['items'][0]['item_code'] == 'VALID'

    def test_numeric_values(self, parser, tmp_path):
        """Test that numeric values are parsed correctly."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['Item Code', 'Item Name', 'Cost Price', 'Sell Price', 'Min Qty', 'Max Qty'])
        ws.append(['TEST', 'Test Item', 50.50, 99.99, 1, 100])
        ws.append(['TEST2', 'Test Item 2', 'invalid', None, 'bad', ''])

        file_path = tmp_path / "numeric.xlsx"
        wb.save(file_path)

        result = parser.parse_inventory_file(str(file_path))

        items = {i['item_code']: i for i in result['items']}

        # Valid numeric values
        assert items['TEST']['cost_price'] == 50.50
        assert items['TEST']['sell_price'] == 99.99
        assert items['TEST']['min_qty'] == 1
        assert items['TEST']['max_qty'] == 100

        # Invalid numeric values should default to 0
        assert items['TEST2']['cost_price'] == 0.0
        assert items['TEST2']['sell_price'] == 0.0
